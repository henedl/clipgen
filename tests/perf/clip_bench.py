"""Repeatable clip-pipeline benchmark: clips, carded clips, and a reel.

The scan side has scan_bench.py; this is the same treatment for the other
half of the product — cutting. Each scenario runs the real CLI against a
deterministic fixture (testsrc video + generated Excel sheet), parses the
`profile |` report, and prints one row per scenario with the ratio the
pool knob is otherwise tuned blind against (pipeline.clip / pool_wall).

Usage (from the repo root):

    uv run python tests/perf/clip_bench.py                     # sweep + table
    uv run python tests/perf/clip_bench.py --save base.json    # snapshot
    uv run python tests/perf/clip_bench.py --compare base.json # A/B deltas
    uv run python tests/perf/clip_bench.py --scenarios clips --runs 2

Fixtures are built on demand in --input: a 120 s testsrc+sine video named
`clipbench_P01.mp4` and `clipbench.xlsx` with 30 one-participant rows
(varied 3-9 s ranges so encodes are not all identical). Each scenario
writes to its own wiped output dir so reservations and manifests never
leak between runs. `--runs N` keeps the fastest run per scenario.

Row reduction is unit-tested in test_clip_bench_parse.py; the rest is a
thin subprocess driver like scan_bench.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from pathlib import Path

from scan_bench import parse_profile

REPO_ROOT = Path(__file__).resolve().parents[2]

DATA_ROWS = 30  # sheet rows with timestamps; row 6 is the first

# Scenario -> extra CLI args. All select the same 20 rows (sheet rows 6-25)
# so clip counts match across scenarios and runs.
SCENARIOS: dict[str, list[str]] = {
    "clips": ["-r", "6-25", "--no-titlecards"],
    "clips-cards": ["-r", "6-25", "--titlecards"],
    "reel": ["-R", "6-25", "--titlecards"],
}


def summarize(profile: dict[str, dict[str, float]]) -> dict[str, float]:
    """Reduce one run's parsed report to the per-scenario comparison row."""

    def get(label: str) -> dict[str, float]:
        return profile.get(label, {"seconds": 0.0, "n": 0})

    clip = get("pipeline.clip")
    pool = get("pipeline.pool_wall")
    wrap = get("titlecard.wrap")
    return {
        "clip_s": clip["seconds"],
        "clips": int(clip["n"]),
        "pool_wall_s": pool["seconds"],
        "parallelism": clip["seconds"] / pool["seconds"] if pool["seconds"] else 0.0,
        "ffmpeg_s": get("ffmpeg.run")["seconds"],
        "ffmpeg_n": int(get("ffmpeg.run")["n"]),
        "ffprobe_n": int(get("ffprobe.run")["n"]),
        "cards_s": wrap["seconds"],
        "cards_copy": int(get("titlecard.copy")["n"]),
        "cards_reencode": int(get("titlecard.reencode")["n"]),
        "peak_rss_mb": profile.get("peak_rss", {}).get("mb", 0.0),
    }


def keep_best(
    best: dict[str, float] | None, summary: dict[str, float]
) -> dict[str, float]:
    """Fastest successful run wins; a failed run (0 clips) never displaces one."""
    if best is None:
        return summary
    if not summary["clips"]:
        return best
    if not best["clips"] or summary["clip_s"] < best["clip_s"]:
        return summary
    return best


def ensure_fixtures(input_dir: Path, duration: int) -> Path:
    """Build the bench video and sheet if they are not already there."""
    input_dir.mkdir(parents=True, exist_ok=True)
    video = input_dir / "clipbench_P01.mp4"
    if not video.is_file():
        subprocess.run(
            [
                "ffmpeg",
                "-y",
                "-loglevel",
                "error",
                "-f",
                "lavfi",
                "-i",
                f"testsrc=duration={duration}:size=1280x720:rate=30",
                "-f",
                "lavfi",
                "-i",
                f"sine=frequency=220:duration={duration}",
                "-pix_fmt",
                "yuv420p",
                "-c:v",
                "libx264",
                "-g",
                "30",
                "-c:a",
                "aac",
                "-shortest",
                str(video),
            ],
            check=True,
        )
    sheet = input_dir / "clipbench.xlsx"
    if not sheet.is_file():
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Observations"
        ws["A1"] = "clipbench"
        ws["F2"] = "ID"
        ws["G2"] = "P01"
        for col, header in enumerate(
            ("Count", "Reported", "Severity", "Category", "Observation", "Summary"), 1
        ):
            ws.cell(5, col, header)
        for r in range(DATA_ROWS):
            # Varied 3-9 s windows so encodes differ; seconds stay <= 57
            # (an MM:SS with SS > 59 is silently dropped at parse).
            start = 2 + (r * 3) % 49
            ws.cell(6 + r, 3, ("Critical", "Serious", "Moderate", "Minor")[r % 4])
            ws.cell(6 + r, 4, "Onboarding")
            ws.cell(6 + r, 5, f"Observation {r}")
            ws.cell(6 + r, 7, f"0:{start:02d}-0:{start + 3 + r % 5:02d}")
        wb.save(sheet)
    return sheet


def run_scenario(
    scenario: str, input_dir: Path, out_root: Path
) -> dict[str, dict[str, float]]:
    """Run one scenario through the CLI into a wiped output dir."""
    out_dir = out_root / f"bench-{scenario}"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    cmd = [
        "uv",
        "run",
        "clipgen.py",
        "-s",
        str(input_dir / "clipbench.xlsx"),
        *SCENARIOS[scenario],
        "-i",
        str(input_dir),
        "-o",
        str(out_dir),
        "--no-input",
        "--profile",
    ]
    proc = subprocess.run(
        cmd, cwd=REPO_ROOT, capture_output=True, text=True, timeout=1800, check=False
    )
    parsed = parse_profile(proc.stdout + proc.stderr)
    if "pipeline.clip" not in parsed:
        print(f"  ! {scenario}: no pipeline.clip in report (run refused?)")
        tail = "\n".join((proc.stdout + proc.stderr).strip().splitlines()[-5:])
        print("    " + tail.replace("\n", "\n    "))
    return parsed


def print_table(
    rows: dict[str, dict[str, float]], baseline: dict[str, dict[str, float]] | None
) -> None:
    delta_hdr = "    Δclip" if baseline else ""
    print(
        f"{'scenario':<14}{'clip':>9}{'n':>4}{'pool':>9}{'par':>6}"
        f"{'ffmpeg':>9}{'cards':>9}{'copy/re':>9}{'rss':>8}{delta_hdr}"
    )
    for scenario, row in rows.items():
        line = (
            f"{scenario:<14}{row['clip_s']:>8.3f}s{row['clips']:>4d}"
            f"{row['pool_wall_s']:>8.3f}s{row['parallelism']:>6.1f}"
            f"{row['ffmpeg_s']:>8.3f}s{row['cards_s']:>8.3f}s"
            f"{row['cards_copy']:>4d}/{row['cards_reencode']:<4d}"
            f"{row['peak_rss_mb']:>6.0f}MB"
        )
        if baseline:
            base = baseline.get(scenario)
            if base and base.get("clip_s"):
                pct = (row["clip_s"] / base["clip_s"] - 1.0) * 100
                line += f"  {pct:>+7.1f}%"
            else:
                line += "  (no base)"
        print(line)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument(
        "--scenarios",
        default=",".join(SCENARIOS),
        help="comma-separated scenario list (default: all)",
    )
    ap.add_argument(
        "--input",
        default="/tmp/clipbench",
        type=Path,
        help="fixture dir; video and sheet are built here if missing",
    )
    ap.add_argument(
        "--output",
        default=None,
        type=Path,
        help="output root (default: <input>/bench-out)",
    )
    ap.add_argument(
        "--duration",
        default=120,
        type=int,
        help="fixture length in seconds (only used when building)",
    )
    ap.add_argument(
        "--runs",
        default=1,
        type=int,
        help="runs per scenario; the fastest (min clip seconds) is kept",
    )
    ap.add_argument("--save", type=Path, help="write results JSON here")
    ap.add_argument("--compare", type=Path, help="baseline JSON to diff against")
    args = ap.parse_args()

    scenarios = [s.strip() for s in args.scenarios.split(",") if s.strip()]
    unknown = [s for s in scenarios if s not in SCENARIOS]
    if unknown:
        print(f"unknown scenarios: {', '.join(unknown)} (know: {', '.join(SCENARIOS)})")
        return 2
    out_root = args.output or (args.input / "bench-out")
    ensure_fixtures(args.input, args.duration)

    baseline = None
    if args.compare:
        baseline = json.loads(args.compare.read_text())["scenarios"]

    rows: dict[str, dict[str, float]] = {}
    for scenario in scenarios:
        best: dict[str, float] | None = None
        for _ in range(max(1, args.runs)):
            summary = summarize(run_scenario(scenario, args.input, out_root))
            best = keep_best(best, summary)
        assert best is not None  # loop runs at least once
        rows[scenario] = best
        print(f"  {scenario}: {best['clips']} clips in {best['pool_wall_s']:.3f}s wall")

    print()
    print_table(rows, baseline)

    if args.save:
        args.save.write_text(
            json.dumps(
                {
                    "meta": {
                        "video": str(args.input / "clipbench_P01.mp4"),
                        "runs": args.runs,
                    },
                    "scenarios": rows,
                },
                indent=2,
            )
        )
        print(f"\nsaved -> {args.save}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

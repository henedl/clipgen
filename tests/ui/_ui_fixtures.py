"""Generate the throwaway project the UI harness points clipgen at.

Nothing here is checked in: the repo ships no sample media, and the maintainer's
``clipgen-test.xlsx`` is gitignored, so a fresh clone has nothing to render. This
module builds a complete miniature study — two ffmpeg ``testsrc`` videos, a
workbook matching ``spreadsheet.build_sheet_context``'s contract, and one seeded
manifest per subsystem — so all six pages come up with real data instead of
their zero-states.

Everything lands under ``.context/ui-check/``. ``.context/`` is the agent scratch
dir the Conductor harness already puts per-worktree files in, so the harness adds
no top-level directory of its own. Whole-subtree coverage is load-bearing —
per-extension rules would let a future ``.zip`` trace or ``.log`` leak into
``git status`` — and Conductor's own ``.git/info/exclude`` entry is local-only
and absent from a plain clone, which is why ``.gitignore`` carries ``.context/``
as well. Do not "simplify" that line away.

Two lifecycles, because the smoke and ``shot.py`` want different things —
:func:`reset_run_dirs` wipes (hermetic, for a pass/fail gate) and
:func:`ensure_run_dirs` seeds only what is missing (cheap, for iteration). The
input dir is cached by both; the ffmpeg encodes are the only slow part.
"""

import json
import shutil
import subprocess
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
ROOT = REPO_ROOT / ".context" / "ui-check"
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
SHOT_DIR = ROOT / "screenshots"
SETTINGS_DIR = ROOT / "settings"
REPORT_PATH = ROOT / "ui-report.json"

STUDY = "clipgen-ui"  # already normalize_study_name-stable: lowercase, no spaces
PARTICIPANTS = ("P01", "P02")
CLIP_SECONDS = 20


class UiUnavailable(RuntimeError):
    """A prerequisite for the UI harness is missing, with install instructions.

    Deliberately not a ``pytest.skip``: ``shot.py`` imports these helpers as a
    plain script, where a raised ``Skipped`` surfaces as a confusing traceback.
    Only ``tests/ui/conftest.py`` translates this into a skip.
    """


# ---- Input fixtures (cached) ----


def require_ffmpeg() -> None:
    """Raise :class:`UiUnavailable` unless both ffmpeg and ffprobe are on PATH.

    ffprobe matters as much as ffmpeg: Composer probes a duration per video part
    and renders no timeline without one.
    """
    missing = [tool for tool in ("ffmpeg", "ffprobe") if shutil.which(tool) is None]
    if missing:
        raise UiUnavailable(
            f"{' and '.join(missing)} not found on PATH — the UI harness needs them "
            "to build its fixture videos.\n"
            "  brew install ffmpeg    (or scripts/install-ffmpeg-ollama.sh)"
        )


def workbook_path() -> Path:
    return INPUT_DIR / f"{STUDY}.xlsx"


def open_workbook() -> tuple[Any, str]:
    """Open the generated workbook. Returns ``(workbook, "")`` or ``(None, reason)``.

    Both entry points must check this *before* handing the workbook to
    ``build_combined_app``: ``server._init_studio_state`` calls ``sys.exit(1)``
    when ``build_sheet_context`` returns ``None`` (server.py:3077), which would
    abort the caller with a bare ``SystemExit`` and an error naming Studio rather
    than the generator that actually drifted.

    A reason string rather than :class:`UiUnavailable`, deliberately: that
    exception means "prerequisite missing, skip", and a drifted fixture is a bug
    to fail on, not a reason to quietly skip. The caller picks the severity.
    """
    import excel_io
    import spreadsheet

    workbook = excel_io.open_excel_workbook(str(workbook_path()))
    if workbook is None:
        return None, f"fixture workbook failed to open: {workbook_path()}"
    if spreadsheet.build_sheet_context(workbook) is None:
        return None, (
            "the generated fixture workbook no longer satisfies "
            "spreadsheet.build_sheet_context — the layout in "
            "tests/ui/_ui_fixtures.py._make_workbook has drifted from "
            "spreadsheet.py's header contract"
        )
    return workbook, ""


def video_path(participant: str) -> Path:
    """Source video path for a participant.

    The ``{study}_{PID}.mp4`` name is the discovery contract
    (``utils.discover_participant_videos``) *and* the fallback
    ``files.get_source_video_filenames`` derives when the sheet's Filename row is
    blank — which is why the generated workbook leaves that row empty.
    """
    return INPUT_DIR / f"{STUDY}_{participant}.mp4"


def ensure_inputs() -> Path:
    """Build the cached input fixtures if absent; return ``INPUT_DIR``.

    Idempotent by existence: a present non-empty video or workbook is reused.
    Delete ``.context/ui-check/input/`` to force a rebuild (~2 s).
    """
    require_ffmpeg()
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    for participant in PARTICIPANTS:
        path = video_path(participant)
        if not path.is_file() or path.stat().st_size == 0:
            _make_testsrc_video(path)
    if not workbook_path().is_file():
        _make_workbook(workbook_path())
    return INPUT_DIR


def _make_testsrc_video(path: Path) -> None:
    """Encode a synthetic clip. Recipe lifted from tests/screenspace/test_scan_pipeline.py."""
    subprocess.run(
        [
            "ffmpeg",
            "-y",
            "-f",
            "lavfi",
            "-i",
            f"testsrc=duration={CLIP_SECONDS}:size=320x240:rate=30",
            "-fps_mode",
            "vfr",
            "-c:v",
            "libx264",
            "-g",
            "30",
            "-pix_fmt",
            "yuv420p",
            str(path),
        ],
        check=True,
        capture_output=True,
    )


# Six rows across four categories and mixed severities, so Overview's Map and
# Metadata tabs get real spread rather than a degenerate single point. Every
# timestamp stays inside the CLIP_SECONDS-long video.
_OBSERVATION_ROWS: tuple[tuple[str, str, str, str, str, str], ...] = (
    # count, reported, severity, category, observation, P01 times, ...
    ("2", "FALSE", "-2", "Onboarding", "Could not find the export button", "0:01-0:04"),
    ("1", "FALSE", "1", "Onboarding", "Misread the progress label", "0:06-0:09"),
    ("3", "FALSE", "-3", "Bug", "Upload silently failed", "0:10-0:13"),
    ("2", "FALSE", "2", "Behavior", "Read the tooltip aloud", "0:14"),
    ("1", "FALSE", "N/A", "Behavior", "Scrolled past the summary", ""),
    ("4", "FALSE", "-1", "Navigation", "Backtracked twice", "0:15-0:18"),
)
_P02_TIMES = ("0:02-0:05", "", "0:07-0:11", "", "0:12-0:16", "0:17")


def _make_workbook(path: Path) -> None:
    """Write the minimal workbook ``spreadsheet.build_sheet_context`` accepts.

    The geometry mirrors the maintainer's real sheet rather than starting at A1,
    so the same header offsets are exercised: ``ID`` at F2 with participant
    columns to its right *on that row*, ``Observation``/``Category`` on row 5,
    and data rows directly below. ``_find_in_data`` scans row-major and takes the
    first exact match, so none of those three labels may appear earlier.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # "Observations" is in config.WORKSHEET_PRIORITY, so the auto-pick selects it
    # deterministically even if a stray default sheet survives.
    sheet.title = "Observations"

    sheet["A1"] = STUDY  # sheet_data[0][0] becomes the study name
    sheet["F1"] = "Baseline time (optional)"  # exercises _detect_baseline_row
    sheet["F2"] = "ID"
    sheet["F3"] = "Metadata"
    sheet["F4"] = "Filename"  # left blank per participant on purpose (see video_path)
    for offset, participant in enumerate(PARTICIPANTS):
        column = 7 + offset  # G, H — strictly right of ID, on the ID row
        sheet.cell(row=2, column=column, value=participant)
        sheet.cell(row=3, column=column, value=f"Participant {offset + 1}")

    headers = ("Count", "Reported", "Severity", "Category", "Observation", "Summary")
    for offset, header in enumerate(headers):
        sheet.cell(row=5, column=1 + offset, value=header)

    for offset, row in enumerate(_OBSERVATION_ROWS):
        count, reported, severity, category, observation, p01 = row
        excel_row = 6 + offset
        for col_offset, value in enumerate(
            (count, reported, severity, category, observation)
        ):
            sheet.cell(row=excel_row, column=1 + col_offset, value=value)
        sheet.cell(row=excel_row, column=7, value=p01)
        sheet.cell(row=excel_row, column=8, value=_P02_TIMES[offset])

    workbook.save(path)
    workbook.close()


# ---- Run dirs (per-run) ----

_MANIFESTS = ("screenspace", "transcripts", "composer", "workflows")


def reset_run_dirs() -> None:
    """Wipe and re-seed every per-run directory.

    The smoke uses this: a run genuinely mutates the output dir (Workflows
    auto-creates a blueprint when the list is empty, Composer persists UI state),
    and a pass/fail gate wants that state hermetic.
    """
    for directory in (OUTPUT_DIR, SHOT_DIR, SETTINGS_DIR):
        shutil.rmtree(directory, ignore_errors=True)
    REPORT_PATH.unlink(missing_ok=True)
    _make_run_dirs()
    for name in _MANIFESTS:
        _SEEDERS[name]()


def ensure_run_dirs() -> None:
    """Seed only what is missing, keeping everything else.

    ``shot.py`` uses this so an agent iterating on one page neither pays the
    reseed cost nor loses state it just poked in through ``--eval``.
    """
    _make_run_dirs()
    for name in _MANIFESTS:
        if not (OUTPUT_DIR / f"{name}_manifest.json").is_file():
            _SEEDERS[name]()


def _make_run_dirs() -> None:
    for directory in (OUTPUT_DIR, SHOT_DIR, SETTINGS_DIR):
        directory.mkdir(parents=True, exist_ok=True)
    if not (SETTINGS_DIR / "start.json").is_file():
        _seed_start_settings()


def _seed_start_settings() -> None:
    """Seed the Start overlay's recent-projects rail.

    Nothing else writes this file, so without it the rail renders "No recent
    projects yet" and neither the entry layout nor the fold-out is reviewable
    in a screenshot. The mix is deliberate: named and unnamed projects, with
    and without a spreadsheet, and one whose output equals its input.
    """
    now = datetime.now(UTC)
    specs = [
        ("Coffee machine study", "recordings", "clips", "Coffee tracker.xlsx"),
        ("", "onboarding-2026", "onboarding-out", "Onboarding notes.xlsx"),
        ("Checkout flow, round 2", "checkout", "checkout", ""),
        ("Kiosk pilot", "kiosk", "kiosk-out", ""),
        ("", "banking-app", "banking-out", "Banking sessions.xlsx"),
        ("Wearables diary study", "wearables", "wearables-out", "Diary.xlsx"),
    ]
    projects: list[dict[str, Any]] = []
    for hours, (name, in_name, out_name, sheet_label) in enumerate(specs):
        sheet = None
        if sheet_label:
            path = str(ROOT / in_name / sheet_label)
            sheet = {
                "type": "excel",
                "id_or_path": path,
                "label": sheet_label,
                "worksheet": "Data",
            }
        projects.append(
            {
                "name": name,
                "input": str(ROOT / in_name),
                "output": str(ROOT / out_name),
                "spreadsheet": sheet,
                "last_opened": (now - timedelta(hours=1 + hours * 7)).isoformat(),
            }
        )
    (SETTINGS_DIR / "start.json").write_text(
        json.dumps({"persist_enabled": True, "recent_projects": projects}, indent=2),
        encoding="utf-8",
    )


def _write(name: str, payload: dict[str, Any]) -> None:
    """Write a manifest by absolute path.

    Not ``utils.save_json_manifest``: that resolves against
    ``config.OUTPUT_DIR``, which would make seeding order-dependent on when the
    caller patched config.
    """
    (OUTPUT_DIR / name).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# Every manifest below must be non-empty. Each subsystem's `_init_*_state` runs a
# guarded persist on startup that DELETES a manifest it considers empty, and an
# empty one also means the page renders its zero-state — which is not what a
# smoke check wants to photograph.


def _seed_screenspace() -> None:
    _write(
        "screenspace_manifest.json",
        {
            "regions": {
                "toolbar": {
                    "x": 0.05,
                    "y": 0.05,
                    "w": 0.40,
                    "h": 0.20,
                    "source_width": 320,
                    "source_height": 240,
                }
            },
            # Empty on purpose: an active task makes the page open an SSE stream,
            # which both defeats any settle heuristic and risks hanging teardown
            # on the server's connection-thread join.
            "tasks": [],
            "events": [
                {
                    "id": "ui-evt-1",
                    "source_video": f"{STUDY}_P01.mp4",
                    "participant": "P01",
                    "detector": "change",
                    "event_type": "change",
                    "time_in": 2.0,
                    "time_out": 5.0,
                    "confidence": 0.82,
                    "metadata": {},
                    "excluded": False,
                    "task_id": "",
                    "region": "toolbar",
                },
                {
                    "id": "ui-evt-2",
                    "source_video": f"{STUDY}_P02.mp4",
                    "participant": "P02",
                    "detector": "change",
                    "event_type": "change",
                    "time_in": 8.0,
                    "time_out": 9.5,
                    "confidence": 0.64,
                    "metadata": {},
                    "excluded": False,
                    "task_id": "",
                    "region": "toolbar",
                },
            ],
            "stashes": [],
            "per_participant": {},
            "pins": {},
        },
    )


def _seed_transcripts() -> None:
    _write(
        "transcripts_manifest.json",
        {
            "source_transcripts": {
                "P01": {
                    "segments": [
                        {
                            "id": "P01:0",
                            "start": 0.0,
                            "end": 2.5,
                            "text": "Where is the export button.",
                        },
                        {
                            "id": "P01:1",
                            "start": 2.5,
                            "end": 6.0,
                            "text": "I expected it in the toolbar.",
                        },
                        {
                            "id": "P01:2",
                            "start": 6.0,
                            "end": 11.0,
                            "text": "Now the upload seems stuck.",
                        },
                    ],
                    "language": "en",
                    "model": "base",
                },
                "P02": {
                    "segments": [
                        {
                            "id": "P02:0",
                            "start": 0.0,
                            "end": 3.0,
                            "text": "This label is confusing.",
                        },
                    ],
                    "language": "en",
                    "model": "base",
                },
            },
            "corrections": [],
            "marks": [],
        },
    )


def _seed_composer() -> None:
    # composer_server never persists on init, so this one is not at risk of
    # deletion — it exists purely so the cuts track renders non-empty.
    # `markerSources` is left out and backfilled from the server's defaults
    # rather than hard-coding its private source list here.
    _write(
        "composer_manifest.json",
        {
            "cuts": [
                {
                    "id": "ui-cut-1",
                    "participant": "P01",
                    "start": 1.0,
                    "end": 4.0,
                    "label": "export hunt",
                    "createdAt": "2026-01-01T00:00:00",
                }
            ],
            "ui": {"followPlayhead": True},
        },
    )


def _seed_workflows() -> None:
    # BUILTIN_STASHES already make the stash library non-empty, but a *user*
    # blueprint needs >=1 node or _is_empty_workflows_manifest deletes the file
    # on init. Seeding one also suppresses the page's auto-POST of an "Untitled".
    _write(
        "workflows_manifest.json",
        {
            "blueprints": [
                {
                    "id": "ui-bp-1",
                    "name": "UI check",
                    "nodes": [
                        {
                            "id": "n1",
                            "type": "video_source",
                            "params": {"participant": "P01"},
                            "position": {"x": 80, "y": 120},
                        }
                    ],
                    "edges": [],
                    "viewport": {"x": 0, "y": 0, "zoom": 1},
                    "trigger": None,
                    "createdAt": "2026-01-01T00:00:00",
                }
            ],
            "stashes": [],
            "runs": [],
        },
    )


_SEEDERS = {
    "screenspace": _seed_screenspace,
    "transcripts": _seed_transcripts,
    "composer": _seed_composer,
    "workflows": _seed_workflows,
}

"""Integration tests for Start-overlay endpoints on the combined Flask app.

Exercises the routes registered directly on ``combined`` (not on a blueprint):

* ``GET  /api/changelog``
* ``GET  /api/dirs``           ``POST /api/dirs``
* ``POST /api/folder-picker``
* ``POST /api/sessions/record``
* ``GET  /api/spreadsheets/google``
* ``POST /api/spreadsheets/google/auth``

We build the live app via :func:`server.build_combined_app` and drive it
through ``test_client``. State globals (``_google_auth``, ``config.INPUT_DIR``,
etc.) are reset around each test so order doesn't matter.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

pytest.importorskip("flask")

import config
import server
import start_settings


@pytest.fixture(scope="module")
def combined_app():
    """The combined Flask app, built once for the module.

    A build costs ~23 ms, almost all of it compiling the six blueprints' ~200
    Werkzeug URL rules; the app object itself carries no per-test state. What it
    *cannot* share is the process-state wiring ``build_combined_app`` also does —
    ``conftest``'s autouse ``_reset_overview_observation_getter`` /
    ``_reset_thinking_agents_getters`` snapshot-and-restore exactly those getters
    around every test, so a once-only build would have its wiring torn down by
    the first test's teardown. Hence the split: routes here, state per test in
    ``app`` below via ``server._init_combined_state``.
    """
    return server.build_combined_app(worksheet=None, default_page="studio")


@pytest.fixture
def app(combined_app, monkeypatch, tmp_path):
    """Combined Flask app with the worksheet unset and dirs pinned to tmp."""
    in_dir = tmp_path / "in"
    out_dir = tmp_path / "out"
    in_dir.mkdir()
    out_dir.mkdir()

    # Pin config dirs so /api/dirs sees a known starting point.
    monkeypatch.setattr(config, "INPUT_DIR", str(in_dir), raising=False)
    monkeypatch.setattr(config, "OUTPUT_DIR", str(out_dir), raising=False)

    # Persist Start settings to tmp so tests don't touch the user's real file.
    monkeypatch.setattr(
        start_settings, "_settings_path", lambda: tmp_path / "start.json"
    )

    # Reset Google-auth state between tests. The Drive listing cache goes with
    # it: the suite runs -n auto with a random seed, so a listing cached by one
    # test would otherwise answer another test's request.
    monkeypatch.setattr(server, "_google_auth", server._GoogleAuthState())
    monkeypatch.setattr(server, "_google_sheet_list_cache", None)

    # /api/spreadsheets/open reassigns these module globals directly (no
    # fixture owns them), so an opened workbook would otherwise outlive the
    # test and feed its participants to anything later in the same worker that
    # reads the live sheet. monkeypatch restores them however the route left them.
    monkeypatch.setattr(server, "_worksheet", None)
    monkeypatch.setattr(server, "_sheet_context", None)
    monkeypatch.setattr(server, "_sheet_payload_cache", None)
    monkeypatch.setattr(server, "_active_sheet_meta", None)

    # Re-run only the state half of the app build. The monkeypatch calls above
    # already reset what this test file drives directly; this re-wires the
    # cross-module getters the autouse conftest fixtures restore after each test.
    server._init_combined_state(worksheet=None)

    return combined_app


@pytest.fixture
def client(app):
    return app.test_client()


# ---------- /api/changelog --------------------------------------------------


def test_changelog_returns_entries(client):
    """Happy-path: real CHANGELOG.md parses into a non-empty entries list."""
    resp = client.get("/api/changelog")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert isinstance(body["entries"], list)
    # The repo ships a populated CHANGELOG.md; if entries is empty, the
    # parser regex drifted from the file format.
    assert body["entries"], "CHANGELOG.md present but no entries parsed"
    first = body["entries"][0]
    assert {"version", "date", "changes"} <= set(first)
    assert {"tool", "kind", "text"} <= set(first["changes"][0])


def test_changelog_handles_missing_file(client, monkeypatch):
    """A missing CHANGELOG.md returns ok with an empty list (warning logged)."""
    import changelog

    monkeypatch.setattr(changelog, "_changelog_path", lambda: Path("/nope/missing.md"))
    resp = client.get("/api/changelog")
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True, "entries": []}


# ---------- /api/dirs (GET + POST) ------------------------------------------


def test_dirs_get_returns_current(client, tmp_path):
    resp = client.get("/api/dirs")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["input"] == str(tmp_path / "in")
    assert body["output"] == str(tmp_path / "out")


def test_dirs_post_updates_both(client, tmp_path):
    new_in = tmp_path / "alt_in"
    new_in.mkdir()
    new_out = tmp_path / "alt_out"  # doesn't exist yet — POST should mkdir
    resp = client.post(
        "/api/dirs",
        data=json.dumps({"input": str(new_in), "output": str(new_out)}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["input"] == str(new_in)
    assert body["output"] == str(new_out)
    assert new_out.is_dir()  # Output dir auto-created


def test_dirs_post_rejects_nonexistent_input(client, tmp_path):
    bogus = tmp_path / "definitely_not_here"
    resp = client.post(
        "/api/dirs",
        data=json.dumps({"input": str(bogus)}),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False
    assert "input" in body["errors"]


def test_dirs_post_rejects_output_under_file(client, tmp_path):
    """If the path points at an existing file (not a dir), mkdir fails."""
    blocker = tmp_path / "blocker"
    blocker.write_text("not a directory")
    resp = client.post(
        "/api/dirs",
        data=json.dumps({"output": str(blocker / "child")}),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False
    assert "output" in body["errors"]


# ---------- /api/folder-picker ---------------------------------------------


def test_folder_picker_returns_chosen_path(client, monkeypatch):
    """The route returns whatever utils.open_native_folder_picker returns."""
    import utils

    monkeypatch.setattr(
        utils, "open_native_folder_picker", lambda initial="": "/picked"
    )
    resp = client.post(
        "/api/folder-picker",
        data=json.dumps({"initial": "/start"}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True, "path": "/picked"}


def test_folder_picker_returns_null_on_cancel(client, monkeypatch):
    import utils

    monkeypatch.setattr(utils, "open_native_folder_picker", lambda initial="": None)
    resp = client.post(
        "/api/folder-picker",
        data=json.dumps({}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True, "path": None}


# ---------- /api/sessions/record -------------------------------------------


def test_sessions_record_happy_path(client, tmp_path):
    payload = {
        "input": str(tmp_path / "in"),
        "output": str(tmp_path / "out"),
        "spreadsheet": {
            "type": "excel",
            "id_or_path": str(tmp_path / "book.xlsx"),
            "label": "book.xlsx",
        },
    }
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(payload),
        content_type="application/json",
    )
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True}
    saved = start_settings.load_start_settings()
    assert saved["recent_projects"], "Session record didn't persist"
    entry = saved["recent_projects"][0]
    assert entry["input"] == payload["input"]
    assert entry["output"] == payload["output"]
    assert entry["spreadsheet"]["type"] == "excel"


def test_sessions_record_no_sheet_path(client, tmp_path):
    """Spreadsheet omitted → still records a session with spreadsheet=None."""
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {"input": str(tmp_path / "in"), "output": str(tmp_path / "out")}
        ),
        content_type="application/json",
    )
    assert resp.status_code == 200
    saved = start_settings.load_start_settings()
    assert saved["recent_projects"][0]["spreadsheet"] is None


def test_sessions_record_stores_the_project_name(client, tmp_path):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {
                "input": str(tmp_path / "in"),
                "output": str(tmp_path / "out"),
                "name": "  Coffee machine study  ",
            }
        ),
        content_type="application/json",
    )
    assert resp.status_code == 200
    saved = start_settings.load_start_settings()
    assert saved["recent_projects"][0]["name"] == "Coffee machine study"


def test_sessions_record_rejects_non_string_name(client, tmp_path):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {"input": str(tmp_path / "in"), "output": str(tmp_path / "out"), "name": 7}
        ),
        content_type="application/json",
    )
    assert resp.status_code == 400
    assert "name" in resp.get_json()["error"]


def test_sessions_record_omitted_name_keeps_the_stored_one(client, tmp_path):
    payload = {"input": str(tmp_path / "in"), "output": str(tmp_path / "out")}
    client.post(
        "/api/sessions/record",
        data=json.dumps({**payload, "name": "Kept"}),
        content_type="application/json",
    )
    client.post(
        "/api/sessions/record",
        data=json.dumps(payload),
        content_type="application/json",
    )
    saved = start_settings.load_start_settings()
    assert saved["recent_projects"][0]["name"] == "Kept"


def test_sessions_record_rejects_non_string_input(client):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps({"input": 42, "output": "/some/where"}),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False
    assert "input" in body["error"]


def test_sessions_record_rejects_bad_spreadsheet_type(client, tmp_path):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {
                "input": str(tmp_path / "in"),
                "output": str(tmp_path / "out"),
                "spreadsheet": {"type": "csv", "id_or_path": "/x"},
            }
        ),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False
    assert "google" in body["error"] and "excel" in body["error"]


def test_sessions_record_rejects_non_dict_spreadsheet(client, tmp_path):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {
                "input": str(tmp_path / "in"),
                "output": str(tmp_path / "out"),
                "spreadsheet": "not-a-dict",
            }
        ),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False


def test_sessions_record_rejects_missing_spreadsheet_id(client, tmp_path):
    resp = client.post(
        "/api/sessions/record",
        data=json.dumps(
            {
                "input": str(tmp_path / "in"),
                "output": str(tmp_path / "out"),
                "spreadsheet": {"type": "google", "id_or_path": "  "},
            }
        ),
        content_type="application/json",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["ok"] is False
    assert "id_or_path" in body["error"]


# ---------- /api/spreadsheets/excel + /worksheets --------------------------


def _write_workbook(path, titles):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = titles[0]
    for t in titles[1:]:
        wb.create_sheet(t)
    wb.save(path)
    wb.close()


def test_excel_list_includes_modified(client, tmp_path):
    """Each .xlsx entry carries a numeric file-modified time for the picker."""
    _write_workbook(tmp_path / "in" / "study.xlsx", ["Sheet1"])
    resp = client.get("/api/spreadsheets/excel")
    body = resp.get_json()
    assert body["ok"] is True
    names = [f["name"] for f in body["files"]]
    assert "study.xlsx" in names
    entry = next(f for f in body["files"] if f["name"] == "study.xlsx")
    assert isinstance(entry["modified"], (int, float))
    assert entry["modified"] > 0


def test_worksheets_route_excel(client, tmp_path):
    """The worksheets route lists an .xlsx's tabs + the priority recommendation."""
    wb_path = tmp_path / "in" / "multi.xlsx"
    _write_workbook(wb_path, ["Intro", "Data", "Extra"])
    resp = client.get(
        "/api/spreadsheets/worksheets",
        query_string={"type": "excel", "id_or_path": str(wb_path)},
    )
    body = resp.get_json()
    assert body["ok"] is True
    assert body["worksheets"] == ["Intro", "Data", "Extra"]
    assert body["recommended"] == "Data"


def test_worksheets_route_rejects_bad_type(client):
    resp = client.get(
        "/api/spreadsheets/worksheets",
        query_string={"type": "bogus", "id_or_path": "x"},
    )
    assert resp.status_code == 400
    assert resp.get_json()["ok"] is False


def test_worksheets_route_google_requires_auth(client):
    resp = client.get(
        "/api/spreadsheets/worksheets",
        query_string={"type": "google", "id_or_path": "My Study"},
    )
    assert resp.get_json()["ok"] is False


def _write_preview_workbook(path, participants, filename_row=None):
    """Write a one-tab .xlsx with *participants* and an optional Filename row.

    ``filename_row`` maps a participant id to its Filename-row override cell.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Study"
    ws["A2"] = "ID"
    ws["D2"] = "Observation"
    ws["E2"] = "Category"
    for i, pid in enumerate(participants):
        ws.cell(row=2, column=2 + i, value=pid)
    if filename_row:
        ws["A3"] = "Filename"
        for i, pid in enumerate(participants):
            override = filename_row.get(pid)
            if override:
                ws.cell(row=3, column=2 + i, value=override)
    wb.save(path)
    wb.close()


def test_preview_reports_expected_filenames_and_disk_status(client, tmp_path):
    """Each participant gets its expected name plus whether it is on disk."""
    wb_path = tmp_path / "in" / "preview.xlsx"
    _write_preview_workbook(wb_path, ["P01", "P02"])
    (tmp_path / "in" / "study_P01.mp4").write_bytes(b"")

    resp = client.get(
        "/api/spreadsheets/preview",
        query_string={
            "type": "excel",
            "id_or_path": str(wb_path),
            "input_dir": str(tmp_path / "in"),
        },
    )
    body = resp.get_json()
    assert body["ok"] is True
    assert body["study"] == "study"
    assert body["participants"] == [
        {
            "id": "P01",
            "filenames": ["study_P01.mp4"],
            "found": True,
            "override": False,
            "override_value": "",
            "sheet_value": "",
        },
        {
            "id": "P02",
            "filenames": ["study_P02.mp4"],
            "found": False,
            "override": False,
            "override_value": "",
            "sheet_value": "",
        },
    ]
    assert body["unmatched"] == []


def test_preview_honours_the_filename_row_override(client, tmp_path):
    """A plus-separated Filename cell previews both parts, in order."""
    wb_path = tmp_path / "in" / "override.xlsx"
    _write_preview_workbook(
        wb_path, ["P01"], filename_row={"P01": "morning.mp4 + afternoon.mp4"}
    )
    (tmp_path / "in" / "morning.mp4").write_bytes(b"")

    body = client.get(
        "/api/spreadsheets/preview",
        query_string={
            "type": "excel",
            "id_or_path": str(wb_path),
            "input_dir": str(tmp_path / "in"),
        },
    ).get_json()
    entry = body["participants"][0]
    assert entry["filenames"] == ["morning.mp4", "afternoon.mp4"]
    assert entry["override"] is True
    assert entry["found"] is False  # afternoon.mp4 is missing


def test_preview_lists_unclaimed_videos_for_the_override_datalist(client, tmp_path):
    """Footage sitting unused in the folder is what an override is likely to want."""
    wb_path = tmp_path / "in" / "preview.xlsx"
    _write_preview_workbook(wb_path, ["P01"])
    (tmp_path / "in" / "study_P01.mp4").write_bytes(b"")
    (tmp_path / "in" / "session two.mp4").write_bytes(b"")

    body = client.get(
        "/api/spreadsheets/preview",
        query_string={
            "type": "excel",
            "id_or_path": str(wb_path),
            "input_dir": str(tmp_path / "in"),
        },
    ).get_json()
    # The claimed file is not offered; the orphan is.
    assert body["unmatched"] == ["session two.mp4"]


def test_preview_applies_a_stored_user_override(client, tmp_path):
    """A saved override wins over the sheet's own Filename row."""
    wb_path = tmp_path / "in" / "override.xlsx"
    _write_preview_workbook(wb_path, ["P01"], filename_row={"P01": "morning.mp4"})
    (tmp_path / "in" / "actually-this-one.mp4").write_bytes(b"")
    start_settings.set_filename_override(
        "excel", str(wb_path), "Data", "P01", "actually-this-one.mp4"
    )

    body = client.get(
        "/api/spreadsheets/preview",
        query_string={
            "type": "excel",
            "id_or_path": str(wb_path),
            "input_dir": str(tmp_path / "in"),
        },
    ).get_json()
    entry = body["participants"][0]
    assert entry["filenames"] == ["actually-this-one.mp4"]
    assert entry["found"] is True
    assert entry["override_value"] == "actually-this-one.mp4"
    assert entry["sheet_value"] == "morning.mp4"  # what Restore falls back to


def test_override_route_persists_and_recomputes_the_row(client, tmp_path):
    (tmp_path / "in" / "recording 3.mp4").write_bytes(b"")

    body = client.post(
        "/api/spreadsheets/preview/override",
        json={
            "type": "excel",
            "id_or_path": str(tmp_path / "in" / "study.xlsx"),
            "worksheet": "Data",
            "participant": "P01",
            "filename": "recording 3.mp4",
            "study": "study",
            "input_dir": str(tmp_path / "in"),
        },
    ).get_json()

    assert body["ok"] is True
    assert body["row"] == {
        "id": "P01",
        "filenames": ["recording 3.mp4"],
        "found": True,
        "override": True,
        "override_value": "recording 3.mp4",
        "sheet_value": "",
    }
    assert start_settings.filename_overrides(
        "excel", str(tmp_path / "in" / "study.xlsx"), "Data"
    ) == {"P01": "recording 3.mp4"}


def test_override_route_clears_back_to_the_sheet_value(client, tmp_path):
    wb = str(tmp_path / "in" / "study.xlsx")
    start_settings.set_filename_override("excel", wb, "Data", "P01", "wrong.mp4")

    body = client.post(
        "/api/spreadsheets/preview/override",
        json={
            "type": "excel",
            "id_or_path": wb,
            "worksheet": "Data",
            "participant": "P01",
            "filename": "",
            "study": "study",
            "sheet_value": "morning.mp4",
            "input_dir": str(tmp_path / "in"),
        },
    ).get_json()

    assert body["row"]["filenames"] == ["morning.mp4"]
    assert body["row"]["override_value"] == ""
    assert start_settings.filename_overrides("excel", wb, "Data") == {}


def test_override_route_refreshes_the_cached_participant_lists(client, tmp_path):
    """Transcripts/Screenspace cache on the input dir's mtime, which an override
    does not move — the edit has to poke that gate or those pages keep serving
    the previous file."""
    import transcripts_server

    in_dir = tmp_path / "in"
    (in_dir / "study_P01.mp4").write_bytes(b"")
    (in_dir / "recording 3.mp4").write_bytes(b"")
    wb_path = in_dir / "cache.xlsx"
    _write_preview_workbook(wb_path, ["P01"])

    client.post(
        "/api/spreadsheets/open", json={"type": "excel", "id_or_path": str(wb_path)}
    )
    transcripts_server._refresh_participants()
    before = transcripts_server._participants[0]["video_paths"]
    assert [Path(p).name for p in before] == ["study_P01.mp4"]

    client.post(
        "/api/spreadsheets/preview/override",
        json={
            "type": "excel",
            "id_or_path": str(wb_path),
            "worksheet": "Data",
            "participant": "P01",
            "filename": "recording 3.mp4",
            "study": "study",
            "input_dir": str(in_dir),
        },
    )
    transcripts_server._refresh_participants()
    after = transcripts_server._participants[0]["video_paths"]
    assert [Path(p).name for p in after] == ["recording 3.mp4"]


def test_override_route_rejects_a_missing_participant(client):
    body = client.post(
        "/api/spreadsheets/preview/override",
        json={"type": "excel", "id_or_path": "/x.xlsx", "filename": "a.mp4"},
    ).get_json()
    assert body["ok"] is False


def test_open_seeds_the_session_filename_overrides(client, tmp_path):
    """Opening a sheet points the whole session at that sheet's overrides."""
    wb_path = tmp_path / "in" / "open.xlsx"
    _write_preview_workbook(wb_path, ["P01"])
    start_settings.set_filename_override(
        "excel", str(wb_path), "Data", "P01", "recording 3.mp4"
    )

    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": str(wb_path)},
    )
    assert resp.get_json()["ok"] is True
    assert config.FILENAME_OVERRIDES == {"P01": "recording 3.mp4"}

    client.post("/api/spreadsheets/close", json={})
    assert config.FILENAME_OVERRIDES == {}


def test_preview_does_not_swap_the_active_sheet(client, tmp_path):
    """Previewing is read-only: no sheet is loaded as a side effect."""
    wb_path = tmp_path / "in" / "preview.xlsx"
    _write_preview_workbook(wb_path, ["P01"])
    assert client.get("/api/status").get_json()["sheet_loaded"] is False

    client.get(
        "/api/spreadsheets/preview",
        query_string={"type": "excel", "id_or_path": str(wb_path)},
    )
    assert client.get("/api/status").get_json()["sheet_loaded"] is False


def test_preview_reports_an_unreadable_worksheet(client, tmp_path):
    """A sheet with no clipgen headers fails with guidance, not a stack trace."""
    import openpyxl

    wb_path = tmp_path / "in" / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "nothing useful"
    wb.active["A2"] = "still nothing"
    wb.save(wb_path)
    wb.close()

    resp = client.get(
        "/api/spreadsheets/preview",
        query_string={"type": "excel", "id_or_path": str(wb_path)},
    )
    body = resp.get_json()
    assert body["ok"] is False
    assert "Observation" in body["error"]


def test_preview_rejects_bad_type(client):
    resp = client.get(
        "/api/spreadsheets/preview",
        query_string={"type": "bogus", "id_or_path": "x"},
    )
    assert resp.status_code == 400
    assert resp.get_json()["ok"] is False


def test_preview_google_requires_auth(client):
    resp = client.get(
        "/api/spreadsheets/preview",
        query_string={"type": "google", "id_or_path": "My Study"},
    )
    assert resp.get_json()["ok"] is False


def _write_valid_workbook(path, tabs):
    """Write an .xlsx whose every tab has a minimal valid clipgen layout."""
    import openpyxl

    wb = openpyxl.Workbook()
    for i, (title, pid) in enumerate(tabs):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        ws["A1"] = "Study"
        ws["A2"] = "ID"
        ws["B2"] = pid
        ws["D2"] = "Observation"
        ws["E2"] = "Category"
        ws["A3"] = "1"
        ws["B3"] = "00:10-00:20"
        ws["D3"] = "Obs"
        ws["E3"] = "CatA"
    wb.save(path)
    wb.close()


def test_open_excel_with_worksheet_reflected_in_status(client, tmp_path):
    """Opening a chosen tab loads it and surfaces it via /api/status."""
    wb_path = tmp_path / "in" / "multi.xlsx"
    _write_valid_workbook(
        wb_path, [("Intro", "P01"), ("Data", "P02"), ("Extra", "P03")]
    )
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": str(wb_path), "worksheet": "Extra"},
    )
    assert resp.get_json()["ok"] is True
    st = client.get("/api/status").get_json()
    assert st["spreadsheet_worksheet"] == "Extra"
    assert st["spreadsheet_label"].endswith("(Extra)")
    # Confirm the Extra tab's data actually loaded (its participant is P03).
    sheet = client.get("/studio/api/sheet").get_json()
    assert sheet["participants"] == ["P03"]


def test_open_excel_records_the_project_name(client, tmp_path):
    """The Start overlay rides project_name along with the sheet open."""
    wb_path = tmp_path / "in" / "named.xlsx"
    _write_valid_workbook(wb_path, [("Data", "P01")])
    resp = client.post(
        "/api/spreadsheets/open",
        json={
            "type": "excel",
            "id_or_path": str(wb_path),
            "project_name": "Coffee machine study",
        },
    )
    assert resp.get_json()["ok"] is True
    saved = start_settings.load_start_settings()
    assert saved["recent_projects"][0]["name"] == "Coffee machine study"


def test_open_excel_rejects_non_string_project_name(client, tmp_path):
    wb_path = tmp_path / "in" / "named.xlsx"
    _write_valid_workbook(wb_path, [("Data", "P01")])
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": str(wb_path), "project_name": 7},
    )
    assert resp.status_code == 400
    assert "project_name" in resp.get_json()["error"]


def test_open_excel_without_worksheet_uses_priority_tab(client, tmp_path):
    """No worksheet field → priority auto-pick ("Data"), not the first tab."""
    wb_path = tmp_path / "in" / "multi.xlsx"
    _write_valid_workbook(
        wb_path, [("Intro", "P01"), ("Data", "P02"), ("Extra", "P03")]
    )
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": str(wb_path)},
    )
    assert resp.get_json()["ok"] is True
    st = client.get("/api/status").get_json()
    assert st["spreadsheet_worksheet"] == "Data"


# ---------- /api/spreadsheets/google + /auth -------------------------------


def test_google_list_unauthenticated(client):
    """No cached client → authenticated=False, no in-flight flag."""
    resp = client.get("/api/spreadsheets/google")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["authenticated"] is False
    assert body["auth_in_flight"] is False
    assert body["sheets"] == []


def test_google_list_unauthenticated_explains_credentials(client):
    """The overlay could say "check credentials.json" but never what that file
    is or where it goes — the searched paths only ever reached stdout, which a
    windowed launch has no console for."""
    import cli

    body = client.get("/api/spreadsheets/google").get_json()
    assert body["credentials_filename"] == "credentials.json"
    assert body["credentials_paths"] == [str(p) for p in cli.credentials_search_paths()]
    assert len(body["credentials_paths"]) >= 2
    assert "gspread" in body["credentials_guide_url"]
    # Empty string (not null) when nothing is on disk, so the JS can branch on
    # truthiness without a null check.
    assert isinstance(body["credentials_found"], str)


def test_google_list_authenticated(client, monkeypatch):
    """Sheet metadata (incl. modifiedTime) comes from get_all_spreadsheet_meta."""
    server._google_auth.client = object()  # stand-in
    import google_api

    monkeypatch.setattr(
        google_api,
        "get_all_spreadsheet_meta",
        lambda _c: [
            {"name": "Alpha", "id": "a", "modifiedTime": "2026-07-20T10:00:00Z"},
            {"name": "Beta", "id": "b", "modifiedTime": ""},
        ],
    )
    resp = client.get("/api/spreadsheets/google")
    body = resp.get_json()
    assert body["authenticated"] is True
    assert [s["name"] for s in body["sheets"]] == ["Alpha", "Beta"]
    # id stays the name (open-by-name contract), modifiedTime flows through.
    assert body["sheets"][0]["id"] == "Alpha"
    assert body["sheets"][0]["modifiedTime"] == "2026-07-20T10:00:00Z"


def test_google_list_surfaces_api_error(client, monkeypatch):
    server._google_auth.client = object()
    import google_api

    def _boom(_c):
        raise RuntimeError("rate limited")

    monkeypatch.setattr(google_api, "get_all_spreadsheet_meta", _boom)
    resp = client.get("/api/spreadsheets/google")
    body = resp.get_json()
    assert body["authenticated"] is True
    assert "rate limited" in body["auth_error"]
    assert body["sheets"] == []


def _count_drive_listings(monkeypatch, metas=None):
    """Patch the Drive listing with a call counter. Returns the counter list."""
    import google_api

    calls: list[int] = []
    rows = metas if metas is not None else [{"name": "Alpha", "id": "a"}]

    def _list(_c):
        calls.append(1)
        return [dict(m) for m in rows]

    monkeypatch.setattr(google_api, "get_all_spreadsheet_meta", _list)
    return calls


def test_google_list_second_call_is_served_from_cache(client, monkeypatch):
    """Within the TTL a repeat listing costs no Drive round-trip."""
    server._google_auth.client = object()
    calls = _count_drive_listings(monkeypatch)

    first = client.get("/api/spreadsheets/google").get_json()
    second = client.get("/api/spreadsheets/google").get_json()

    assert len(calls) == 1
    assert first["sheets"] == second["sheets"]


def test_google_list_refresh_param_relists(client, monkeypatch):
    """?refresh=true is the picker's escape hatch for a mid-session sheet."""
    server._google_auth.client = object()
    calls = _count_drive_listings(monkeypatch)

    client.get("/api/spreadsheets/google")
    client.get("/api/spreadsheets/google", query_string={"refresh": "true"})

    assert len(calls) == 2


def test_google_list_cache_expires(client, monkeypatch):
    server._google_auth.client = object()
    calls = _count_drive_listings(monkeypatch)
    monkeypatch.setattr(server, "_GOOGLE_SHEET_LIST_TTL_SEC", 0)

    client.get("/api/spreadsheets/google")
    client.get("/api/spreadsheets/google")

    assert len(calls) == 2


def test_google_list_failure_is_not_cached(client, monkeypatch):
    """A rate-limited listing must not poison the cache with an empty result."""
    server._google_auth.client = object()
    import google_api

    def _boom(_c):
        raise RuntimeError("rate limited")

    monkeypatch.setattr(google_api, "get_all_spreadsheet_meta", _boom)
    assert (
        "rate limited"
        in client.get("/api/spreadsheets/google").get_json()["auth_error"]
    )

    _count_drive_listings(monkeypatch)
    body = client.get("/api/spreadsheets/google").get_json()
    assert [s["name"] for s in body["sheets"]] == ["Alpha"]


def test_google_auth_invalidates_the_listing_cache(client, monkeypatch):
    """A newly signed-in account must not inherit the previous one's listing."""
    import threading

    server._google_sheet_list_cache = (1.0, [{"name": "Stale", "id": "s"}])

    def fake_thread(target, daemon=True):
        target()

        class _Joined:
            def start(self_inner):
                pass

        return _Joined()

    monkeypatch.setattr(threading, "Thread", fake_thread)
    import cli

    monkeypatch.setattr(cli, "authenticate_google", lambda: object())

    client.post("/api/spreadsheets/google/auth", json={})
    assert server._google_sheet_list_cache is None


def test_open_by_name_relists_when_the_name_is_missing(monkeypatch):
    """A sheet created after the cache filled stays openable inside the TTL."""
    import app as _app

    monkeypatch.setattr(server, "_google_auth", server._GoogleAuthState())
    server._google_auth.client = object()
    monkeypatch.setattr(server, "_google_sheet_list_cache", None)
    calls = _count_drive_listings(monkeypatch, [{"name": "Alpha", "id": "a"}])

    seen: list[list[str]] = []

    def _open(_client, doc_list, _name, worksheet_name=None):
        seen.append(list(doc_list))
        return object()

    monkeypatch.setattr(_app, "open_spreadsheet_by_name", _open)

    server._open_worksheet_for("google", "Alpha", None)
    assert len(calls) == 1  # cached name: no second listing

    server._open_worksheet_for("google", "Brand New", None)
    assert len(calls) == 2  # unknown name: one forced re-list
    assert seen == [["Alpha"], ["Alpha"]]


def test_worksheets_route_reuses_the_cached_names(client, monkeypatch):
    """The worksheet dropdown takes the cached doc list instead of re-listing."""
    import app as _app

    server._google_auth.client = object()
    calls = _count_drive_listings(monkeypatch)

    received: list[list[str] | None] = []

    def _titles(_client, _id_or_path, doc_list=None):
        received.append(doc_list)
        return ["Data"], "Data"

    monkeypatch.setattr(_app, "list_worksheet_titles", _titles)

    client.get("/api/spreadsheets/google")  # primes the cache
    resp = client.get(
        "/api/spreadsheets/worksheets",
        query_string={"type": "google", "id_or_path": "Alpha"},
    )

    assert resp.get_json()["worksheets"] == ["Data"]
    assert received == [["Alpha"]]
    assert len(calls) == 1


def test_worksheets_route_google_url_skips_the_drive_listing(client, monkeypatch):
    """A URL resolves without a listing — the doc list stays None."""
    import app as _app

    server._google_auth.client = object()
    calls = _count_drive_listings(monkeypatch)

    received: list[list[str] | None] = []

    def _titles(_client, _id_or_path, doc_list=None):
        received.append(doc_list)
        return ["Data"], "Data"

    monkeypatch.setattr(_app, "list_worksheet_titles", _titles)
    client.get(
        "/api/spreadsheets/worksheets",
        query_string={
            "type": "google",
            "id_or_path": "https://docs.google.com/spreadsheets/d/abc",
        },
    )

    assert received == [None]
    assert calls == []


def test_google_auth_short_circuits_when_already_authenticated(client):
    """POST /auth returns authenticated=True without spawning a thread."""
    server._google_auth.client = object()
    resp = client.post("/api/spreadsheets/google/auth", json={})
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["started"] is False
    assert body["authenticated"] is True


def test_google_auth_short_circuits_when_in_flight(client):
    server._google_auth.in_flight = True
    resp = client.post("/api/spreadsheets/google/auth", json={})
    body = resp.get_json()
    assert body["started"] is False
    assert body["in_flight"] is True


def test_google_auth_runs_handler_when_idle(client, monkeypatch):
    """When no client and no flight, the thread runs cli.authenticate_google."""
    import threading

    # Force the spawned thread to run inline so we can assert end state.
    def fake_thread(target, daemon=True):
        target()

        class _Joined:
            def start(self_inner):
                pass

        return _Joined()

    monkeypatch.setattr(threading, "Thread", fake_thread)

    sentinel = object()
    import cli

    monkeypatch.setattr(cli, "authenticate_google", lambda: sentinel)

    resp = client.post("/api/spreadsheets/google/auth", json={})
    assert resp.status_code == 202
    body = resp.get_json()
    assert body["started"] is True
    # After the (synchronous) thread ran, state should reflect success.
    assert server._google_auth.client is sentinel
    assert server._google_auth.in_flight is False
    assert server._google_auth.error == ""


def test_google_auth_records_thread_error(client, monkeypatch):
    import threading

    def fake_thread(target, daemon=True):
        target()

        class _Joined:
            def start(self_inner):
                pass

        return _Joined()

    monkeypatch.setattr(threading, "Thread", fake_thread)

    import cli

    def _explode():
        raise RuntimeError("creds missing")

    monkeypatch.setattr(cli, "authenticate_google", _explode)

    client.post("/api/spreadsheets/google/auth", json={})
    assert server._google_auth.client is None
    assert "creds missing" in server._google_auth.error
    assert server._google_auth.in_flight is False


# ---------- sheet-switch guard during active generation ---------------------


def test_spreadsheets_open_rejected_during_generation(client, monkeypatch):
    """Switching spreadsheets is rejected with 409 while a clip generation is
    in progress, so the generated lists are not rebound under an active stream."""
    monkeypatch.setitem(server._busy_slots, "generate", True)
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": "/tmp/whatever.xlsx"},
    )
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["ok"] is False
    assert "in progress" in data["error"]


def test_spreadsheets_open_rejected_during_intake(client, monkeypatch):
    """An in-flight intake stream also blocks a spreadsheet switch."""
    monkeypatch.setattr(server, "_intake_active", 1)
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": "/tmp/whatever.xlsx"},
    )
    assert resp.status_code == 409
    assert resp.get_json()["ok"] is False


def test_spreadsheets_open_rejected_during_timeline_viewer(client, monkeypatch):
    """A timeline-viewer build blocks a spreadsheet switch: it appends into the
    shared generated list/manifest, so a swap mid-build would rebind those under
    it and mix old-sheet artifacts into the new sheet."""
    monkeypatch.setitem(server._busy_slots, "timeline_viewer", True)
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": "/tmp/whatever.xlsx"},
    )
    assert resp.status_code == 409
    assert resp.get_json()["ok"] is False


def test_spreadsheets_open_rejected_during_gallery(client, monkeypatch):
    """A gallery build also blocks a spreadsheet switch."""
    monkeypatch.setitem(server._busy_slots, "gallery", True)
    resp = client.post(
        "/api/spreadsheets/open",
        json={"type": "excel", "id_or_path": "/tmp/whatever.xlsx"},
    )
    assert resp.status_code == 409
    assert resp.get_json()["ok"] is False


def test_spreadsheets_close_rejected_during_reel(client, monkeypatch):
    """Closing the spreadsheet is rejected with 409 while a reel build runs."""
    monkeypatch.setitem(server._busy_slots, "reel", True)
    resp = client.post("/api/spreadsheets/close")
    assert resp.status_code == 409
    assert resp.get_json()["ok"] is False


def test_combined_server_forces_non_interactive(monkeypatch):
    """The web server has no console: serve_combined_app must flip
    utils.NO_INPUT_MODE on so a missing source video is skipped-and-reported
    instead of blocking a Flask/daemon thread on input() (regression: watch-dir
    triggered runs + Studio generate hung on the fuzzy-match prompt)."""
    import utils

    def _fake_app(environ, start_response):
        start_response("200 OK", [("Content-Type", "text/plain")])
        return [b"ok"]

    monkeypatch.setattr(utils, "NO_INPUT_MODE", False)
    monkeypatch.setattr(utils, "preload_vision_libs_quietly", lambda **kwargs: None)
    monkeypatch.setattr(utils, "sweep_stale_temp_artifacts", lambda: None)
    monkeypatch.setattr(server, "build_combined_app", lambda **kwargs: _fake_app)

    live = server.serve_combined_app(port=0, block_until_ready=True)
    try:
        assert utils.NO_INPUT_MODE is True
        assert live.boot["ready"] is True
    finally:
        # Close the socket directly: stop_combined_app would import the heavy
        # blueprint modules to stop workers the fake app never started.
        live.srv.shutdown()
        live.srv.server_close()


def test_start_settings_toggles_are_independent(client):
    """Turning off project history must not also stop remembering the window.

    They are two checkboxes with two meanings; wiring the window rect to
    persist_enabled made the feature a silent no-op for anyone who had already
    opted out of recents.
    """
    resp = client.post(
        "/api/start-settings",
        data=json.dumps({"persist_enabled": False}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    settings = resp.get_json()["settings"]
    assert settings["persist_enabled"] is False
    assert settings["remember_window"] is True

    start_settings.record_window_geometry(10, 20, 1200, 800)
    assert start_settings.load_window_geometry() is not None


def test_start_settings_remember_window_toggle(client):
    resp = client.post(
        "/api/start-settings",
        data=json.dumps({"remember_window": False}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    assert resp.get_json()["settings"]["remember_window"] is False

    start_settings.record_window_geometry(10, 20, 1200, 800)
    assert start_settings.load_window_geometry() is None


def test_start_settings_get_reports_desktop_launch(client, monkeypatch):
    """The overlay only offers the window toggle when there is a window."""
    import utils

    monkeypatch.setattr(utils, "GUI_LAUNCH", False)
    assert client.get("/api/start-settings").get_json()["desktop"] is False
    monkeypatch.setattr(utils, "GUI_LAUNCH", True)
    assert client.get("/api/start-settings").get_json()["desktop"] is True


# ---------- startup notice ---------------------------------------------------


def test_status_startup_notice_empty_by_default(client):
    assert client.get("/api/status").get_json()["startup_notice"] == ""


def test_status_carries_startup_notice(client, monkeypatch):
    monkeypatch.setattr(
        server,
        "_startup_notice",
        {"message": "boot could not open 'X'", "source_type": "excel"},
    )
    s = client.get("/api/status").get_json()
    assert s["startup_notice"] == "boot could not open 'X'"
    assert s["startup_notice_source"] == "excel"


def test_successful_open_clears_startup_notice(client, monkeypatch, tmp_path):
    """Opening any sheet moots whatever the boot build failed to open."""
    monkeypatch.setattr(
        server,
        "_startup_notice",
        {"message": "boot could not open 'X'", "source_type": "google"},
    )
    wb_path = tmp_path / "in" / "study.xlsx"
    _write_preview_workbook(wb_path, ["P01"])
    resp = client.post(
        "/api/spreadsheets/open", json={"type": "excel", "id_or_path": str(wb_path)}
    )
    assert resp.get_json()["ok"] is True
    assert client.get("/api/status").get_json()["startup_notice"] == ""


# ---------- /api/settings ---------------------------------------------------


def test_settings_get_reports_the_config_dir_path(client, monkeypatch, tmp_path):
    """The GET carries the settings path so the modal can show it."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    body = client.get("/api/settings").get_json()
    assert body["ok"] is True
    assert body["path"] == str(tmp_path / config.STUDIO_SETTINGS_FILENAME)


@pytest.mark.parametrize("gui_launch", [True, False])
def test_settings_get_reports_the_desktop_flag(client, monkeypatch, gui_launch):
    """The reveal button follows GUI_LAUNCH, so every native window gets it.

    Not ``html[data-desktop-chrome]``: that is macOS-only, and a Windows or
    Linux webview has no address bar either.
    """
    monkeypatch.setattr(server.utils, "GUI_LAUNCH", gui_launch)
    assert client.get("/api/settings").get_json()["desktop"] is gui_launch


def test_settings_put_writes_beside_start_json(client, monkeypatch, tmp_path):
    """A settings PUT lands in the config dir, never in the output dir."""
    cfg = tmp_path / "cfg"
    monkeypatch.setattr(start_settings, "config_dir", lambda: cfg)
    monkeypatch.setattr(server.config, "WEBP_QUALITY", server.config.WEBP_QUALITY)

    resp = client.put("/api/settings", json={"settings": {"WEBP_QUALITY": 55}})
    assert resp.get_json()["ok"] is True

    saved = json.loads(
        (cfg / config.STUDIO_SETTINGS_FILENAME).read_text(encoding="utf-8")
    )
    assert saved["WEBP_QUALITY"] == 55
    output_copy = Path(config.OUTPUT_DIR) / config.STUDIO_SETTINGS_FILENAME
    assert not output_copy.exists()


def test_settings_reveal_shows_the_file(client, monkeypatch, tmp_path):
    """The reveal route hands the settings file to the OS file browser."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    settings_file = tmp_path / config.STUDIO_SETTINGS_FILENAME
    settings_file.write_text("{}", encoding="utf-8")
    shown: list[Path] = []
    monkeypatch.setattr(
        server.utils, "reveal_in_file_manager", lambda p: shown.append(p) or True
    )

    body = client.post("/api/settings/reveal", json={}).get_json()
    assert body["ok"] is True
    assert shown == [settings_file]


def test_settings_reveal_falls_back_to_the_folder(client, monkeypatch, tmp_path):
    """With everything at default there is no file, so the folder opens."""
    cfg = tmp_path / "cfg"
    monkeypatch.setattr(start_settings, "config_dir", lambda: cfg)
    shown: list[Path] = []
    monkeypatch.setattr(
        server.utils, "reveal_in_file_manager", lambda p: shown.append(p) or True
    )

    body = client.post("/api/settings/reveal", json={}).get_json()
    assert body["ok"] is True
    assert shown == [cfg]
    assert cfg.is_dir()


def test_settings_reveal_reports_failure(client, monkeypatch, tmp_path):
    """A file browser that would not start is an error, not a silent ok."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    (tmp_path / config.STUDIO_SETTINGS_FILENAME).write_text("{}", encoding="utf-8")
    monkeypatch.setattr(server.utils, "reveal_in_file_manager", lambda p: False)

    body = client.post("/api/settings/reveal", json={}).get_json()
    assert body["ok"] is False
    assert "folder" in body["error"]


# ---------- /api/models/llm ------------------------------------------------


def test_llm_reveal_shows_the_gguf(client, monkeypatch, tmp_path):
    """Reveal hands the model's file in the models dir to the file browser."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    models = tmp_path / "models"
    models.mkdir()
    gguf = models / "tiny.gguf"
    gguf.write_bytes(b"stub")
    shown: list[Path] = []
    monkeypatch.setattr(
        server.utils, "reveal_in_file_manager", lambda p: shown.append(p) or True
    )

    body = client.post("/api/models/llm/reveal", json={"model": "tiny"}).get_json()
    assert body["ok"] is True
    assert shown == [gguf]
    assert body["path"] == str(gguf)


def test_llm_reveal_rejects_an_unknown_model(client, monkeypatch, tmp_path):
    """Nothing on disk under that name is a 404, not a blank file-browser call."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    monkeypatch.setattr(server.utils, "reveal_in_file_manager", lambda p: True)

    resp = client.post("/api/models/llm/reveal", json={"model": "ghost"})
    assert resp.status_code == 404
    assert resp.get_json()["error"] == "Model not found"


def test_llm_reveal_reports_failure(client, monkeypatch, tmp_path):
    """A file browser that would not start is an error, not a silent ok."""
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    models = tmp_path / "models"
    models.mkdir()
    (models / "tiny.gguf").write_bytes(b"stub")
    monkeypatch.setattr(server.utils, "reveal_in_file_manager", lambda p: False)

    body = client.post("/api/models/llm/reveal", json={"model": "tiny"}).get_json()
    assert body["ok"] is False
    assert "folder" in body["error"]


def test_llm_delete_is_reachable_from_the_combined_root(client, monkeypatch, tmp_path):
    """The settings modal opens from every page, so it calls the root path.

    The rule itself lives on the transcripts blueprint; without the root
    registration this 404s on routing and the Delete button does nothing.
    """
    monkeypatch.setattr(start_settings, "config_dir", lambda: tmp_path)
    models = tmp_path / "models"
    models.mkdir()
    gguf = models / "tiny.gguf"
    gguf.write_bytes(b"stub")

    body = client.delete("/api/models/llm/tiny").get_json()
    assert body["ok"] is True
    assert not gguf.exists()

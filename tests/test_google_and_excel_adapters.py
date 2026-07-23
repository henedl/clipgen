from typing import Any, cast

import gspread
import google_api
import excel_io
import utils


class _FakeWorksheet:
    def __init__(self, title):
        self.title = title


class _FakeSpreadsheet:
    def __init__(self, titles):
        self._worksheets = [_FakeWorksheet(t) for t in titles]

    def worksheets(self):
        return self._worksheets

    def worksheet(self, title):
        for ws in self._worksheets:
            if ws.title == title:
                return ws
        raise KeyError(title)


def test_pick_worksheet_title_precedence():
    titles = ["Intro", "Data", "Extra"]
    # Explicit preferred name wins when present.
    assert utils.pick_worksheet_title(titles, "Extra") == "Extra"
    # A missing preferred name falls back to the priority match ("Data").
    assert utils.pick_worksheet_title(titles, "Nope") == "Data"
    # No preferred name → priority match.
    assert utils.pick_worksheet_title(titles) == "Data"
    # No priority match → first title.
    assert utils.pick_worksheet_title(["Alpha", "Beta"]) == "Alpha"
    # Empty list → None.
    assert utils.pick_worksheet_title([]) is None


def test_get_worksheet_prefers_priority_match(monkeypatch):
    fake_spreadsheet = _FakeSpreadsheet(["Other", "Sheet1"])
    ws = google_api.get_worksheet(cast(gspread.Spreadsheet, fake_spreadsheet))
    assert ws.title == "Sheet1"


def test_get_worksheet_honors_preferred_name():
    fake_spreadsheet = _FakeSpreadsheet(["Other", "Sheet1", "Extra"])
    ws = google_api.get_worksheet(
        cast(gspread.Spreadsheet, fake_spreadsheet), preferred_name="Extra"
    )
    assert ws.title == "Extra"


def test_get_worksheet_falls_back_when_preferred_absent():
    # Requested tab is gone → priority auto-pick ("Sheet1"), never an error.
    fake_spreadsheet = _FakeSpreadsheet(["Other", "Sheet1"])
    ws = google_api.get_worksheet(
        cast(gspread.Spreadsheet, fake_spreadsheet), preferred_name="Ghost"
    )
    assert ws.title == "Sheet1"


def test_get_all_spreadsheet_meta_passes_through_modified_time():
    class FakeClient:
        def list_spreadsheet_files(self):
            return [
                {"id": "a", "name": "Alpha", "modifiedTime": "2026-07-20T10:00:00Z"},
                {"id": "b", "name": "Beta"},  # modifiedTime absent → ""
            ]

    metas = google_api.get_all_spreadsheet_meta(cast(gspread.Client, FakeClient()))
    assert metas == [
        {"id": "a", "name": "Alpha", "modifiedTime": "2026-07-20T10:00:00Z"},
        {"id": "b", "name": "Beta", "modifiedTime": ""},
    ]
    # The name-only wrapper stays compatible for CLI callers.
    assert google_api.get_all_spreadsheets(cast(gspread.Client, FakeClient())) == [
        "Alpha",
        "Beta",
    ]


def test_find_spreadsheet_by_name_and_guess(monkeypatch):
    monkeypatch.setattr(google_api.config, "DEBUGGING", False, raising=False)
    docs = ["Playtest Data Set", "Other Sheet"]

    idx_exact = google_api.find_spreadsheet_by_name("Other Sheet", docs)
    assert idx_exact == 1

    idx_guess = google_api.find_spreadsheet_by_name("Playtest", docs)
    assert idx_guess == 0


def test_find_spreadsheet_by_name_prefers_exact_over_guess(monkeypatch):
    monkeypatch.setattr(google_api.config, "DEBUGGING", False, raising=False)
    docs = ["Playtest Data Set", "Playtest"]

    idx = google_api.find_spreadsheet_by_name("Playtest", docs)
    assert idx == 1

    docs_reversed = ["Playtest", "Playtest Data Set"]
    idx_reversed = google_api.find_spreadsheet_by_name("Playtest", docs_reversed)
    assert idx_reversed == 0


def test_get_sheet_values_retries_on_rate_limit(monkeypatch):
    class FakeAPIError(gspread.exceptions.APIError):
        def __init__(self, status_code: int):
            self.response = cast(Any, type("R", (), {"status_code": status_code})())

        def __str__(self) -> str:
            return f"HTTP {self.response.status_code}"

    calls = {"n": 0}

    class FakeSheet:
        def get_all_values(self):
            calls["n"] += 1
            if calls["n"] == 1:
                raise FakeAPIError(429)
            return [["Study"], ["ID", "P01"]]

    monkeypatch.setattr(google_api.time, "sleep", lambda _s: None)
    assert google_api.get_sheet_values(FakeSheet()) == [
        ["Study"],
        ["ID", "P01"],
    ]
    assert calls["n"] == 2


def test_get_sheet_values_does_not_retry_client_errors():
    class FakeAPIError(gspread.exceptions.APIError):
        def __init__(self, status_code: int):
            self.response = cast(Any, type("R", (), {"status_code": status_code})())

        def __str__(self) -> str:
            return f"HTTP {self.response.status_code}"

    calls = {"n": 0}

    class FakeSheet:
        def get_all_values(self):
            calls["n"] += 1
            raise FakeAPIError(403)

    import pytest

    with pytest.raises(FakeAPIError):
        google_api.get_sheet_values(FakeSheet())
    assert calls["n"] == 1


def test_excel_sheet_adapter_basic_access(tmp_path, monkeypatch):
    # Build a tiny workbook with one sheet and some data.
    import openpyxl

    wb_path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Study"
    ws["A2"] = "ID"
    ws["B2"] = "P01"
    ws["D2"] = "Observation"
    ws["E2"] = "Category"
    ws["A3"] = "1"
    ws["B3"] = "00:10-00:20"
    ws["D3"] = "Obs one"
    ws["E3"] = "CatA"
    wb.save(wb_path)
    wb.close()

    adapter = excel_io.open_excel_workbook(str(wb_path))
    assert adapter is not None

    # Adapter should expose spreadsheet-like metadata.
    assert adapter.spreadsheet.title == "test"
    assert adapter.spreadsheet.url is None

    # Data access helpers.
    all_values = adapter.get_all_values()
    assert all_values[0][0] == "Study"

    id_cell = adapter.find("ID")
    assert id_cell is not None
    assert id_cell.row == 2

    row2 = adapter.row_values(2)
    assert "P01" in row2
    assert adapter.col_count >= 5


def _make_multi_tab_workbook(path):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Intro"
    wb.create_sheet("Data")
    wb.create_sheet("Extra")
    wb.save(path)
    wb.close()


def test_list_worksheet_titles_returns_titles_and_recommended(tmp_path):
    wb_path = tmp_path / "multi.xlsx"
    _make_multi_tab_workbook(wb_path)
    titles, recommended = excel_io.list_worksheet_titles(str(wb_path))
    assert titles == ["Intro", "Data", "Extra"]
    assert recommended == "Data"  # priority auto-pick


def test_open_excel_workbook_honors_worksheet_name(tmp_path):
    wb_path = tmp_path / "multi.xlsx"
    _make_multi_tab_workbook(wb_path)
    adapter = excel_io.open_excel_workbook(str(wb_path), worksheet_name="Extra")
    assert adapter is not None
    assert adapter.title == "Extra"


def test_open_excel_workbook_falls_back_when_worksheet_absent(tmp_path):
    wb_path = tmp_path / "multi.xlsx"
    _make_multi_tab_workbook(wb_path)
    # Unknown tab → priority auto-pick ("Data"), not an error.
    adapter = excel_io.open_excel_workbook(str(wb_path), worksheet_name="Ghost")
    assert adapter is not None
    assert adapter.title == "Data"


def test_open_excel_workbook_corrupt_file_returns_none(tmp_path, capsys):
    # A file with an .xlsx suffix but non-zip content makes openpyxl raise
    # zipfile.BadZipFile, which does not subclass OSError. The adapter must
    # still fail gracefully (error message + None) rather than propagate.
    bad_path = tmp_path / "corrupt.xlsx"
    bad_path.write_text("this is not a real xlsx file")

    assert excel_io.open_excel_workbook(str(bad_path)) is None


def test_open_excel_workbook_unsupported_suffix_returns_none(tmp_path):
    # A real file openpyxl cannot read by extension raises InvalidFileException
    # (also not an OSError subclass).
    txt_path = tmp_path / "notes.txt"
    txt_path.write_text("plain text, not a spreadsheet")

    assert excel_io.open_excel_workbook(str(txt_path)) is None


def _make_workbook(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ID"
    wb.save(path)
    wb.close()


def test_prompt_for_excel_fallback_picks_by_index(tmp_path, monkeypatch):
    wb_path = tmp_path / "alpha.xlsx"
    _make_workbook(wb_path)
    monkeypatch.chdir(tmp_path)

    inputs = iter(["1"])
    monkeypatch.setattr(excel_io.utils, "read_user_input", lambda _prompt: next(inputs))

    adapter = excel_io.prompt_for_excel_fallback()
    assert adapter is not None
    assert adapter.spreadsheet.title == "alpha"


def test_prompt_for_excel_fallback_accepts_path(tmp_path, monkeypatch):
    # No .xlsx in cwd — user pastes an absolute path.
    cwd = tmp_path / "empty"
    cwd.mkdir()
    elsewhere = tmp_path / "data" / "other.xlsx"
    elsewhere.parent.mkdir()
    _make_workbook(elsewhere)
    monkeypatch.chdir(cwd)

    inputs = iter([str(elsewhere)])
    monkeypatch.setattr(excel_io.utils, "read_user_input", lambda _prompt: next(inputs))

    adapter = excel_io.prompt_for_excel_fallback()
    assert adapter is not None
    assert adapter.spreadsheet.title == "other"


def test_prompt_for_excel_fallback_cancel_returns_none(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    inputs = iter([""])
    monkeypatch.setattr(excel_io.utils, "read_user_input", lambda _prompt: next(inputs))

    assert excel_io.prompt_for_excel_fallback() is None


def test_prompt_for_excel_fallback_help_then_cancel(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    inputs = iter(["help", ""])
    monkeypatch.setattr(excel_io.utils, "read_user_input", lambda _prompt: next(inputs))

    # 'help' should not exit the loop; subsequent empty input cancels.
    assert excel_io.prompt_for_excel_fallback() is None

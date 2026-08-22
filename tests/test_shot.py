"""Argparse and config overrides for tests/ui/shot.py --perf benches.

The UI harness itself is opt-in (``CLIPGEN_UI_CHECK=1``); these tests pin the
plumbing that lets ``--sheet`` / ``--output`` point at a real-sized project
without booting Chromium. ``tests/ui`` is not on the type-checker's path
(norecursedirs), so the helpers load via importlib rather than a bare import.
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from types import ModuleType

import config

UI = Path(__file__).resolve().parent / "ui"


def _load_ui(name: str) -> ModuleType:
    if str(UI) not in sys.path:
        sys.path.insert(0, str(UI))
    path = UI / f"{name}.py"
    spec = importlib.util.spec_from_file_location(f"clipgen_ui_{name}", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_shot_parses_sheet_input_output():
    shot = _load_ui("shot")
    args = shot._parse_args(
        [
            "studio",
            "--sheet",
            "/tmp/gridbench.xlsx",
            "--input",
            "/tmp/ssbench",
            "--output",
            "/tmp/tsbench",
            "--perf",
        ]
    )
    assert args.page == "studio"
    assert args.sheet == Path("/tmp/gridbench.xlsx")
    assert args.input_dir == Path("/tmp/ssbench")
    assert args.output_dir == Path("/tmp/tsbench")
    assert args.perf is True


def test_redirect_config_honors_dir_overrides(monkeypatch, tmp_path):
    import start_settings
    import utils

    session = _load_ui("_ui_session")
    monkeypatch.setattr(config, "INPUT_DIR", "orig-in")
    monkeypatch.setattr(config, "OUTPUT_DIR", "orig-out")
    # redirect_config rebinds these with plain assignment (its callers are
    # standalone scripts that exit afterwards). Pre-register the originals so
    # monkeypatch restores them — without this, the leaked _settings_path
    # lambda fails test_start_settings' path tests when they run later.
    monkeypatch.setattr(config, "VERBOSITY", config.VERBOSITY)
    monkeypatch.setattr(utils, "NO_INPUT_MODE", utils.NO_INPUT_MODE)
    monkeypatch.setattr(start_settings, "_settings_path", start_settings._settings_path)
    monkeypatch.setattr(start_settings, "config_dir", start_settings.config_dir)
    session.redirect_config(input_dir=tmp_path / "in", output_dir=tmp_path / "out")
    assert config.INPUT_DIR == str(tmp_path / "in")
    assert config.OUTPUT_DIR == str(tmp_path / "out")


def _write_gridbench(path: Path, *, rows: int = 3, participants: int = 2) -> None:
    """The profile-skill gridbench geometry, shrunk for a unit test."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observations"
    ws["A1"] = "gridbench"
    ws["F2"] = "ID"
    for i in range(participants):
        ws.cell(2, 7 + i, f"P{i + 1:02d}")
    for col, h in enumerate(
        ("Count", "Reported", "Severity", "Category", "Observation", "Summary"), 1
    ):
        ws.cell(5, col, h)
    sevs = ("Critical", "Serious", "Moderate", "Minor")
    for r in range(rows):
        ws.cell(6 + r, 3, sevs[r % 4])
        ws.cell(6 + r, 4, "Onboarding")
        ws.cell(6 + r, 5, f"Observation {r}")
        for i in range(participants):
            ws.cell(6 + r, 7 + i, "0:01-0:04" if i % 3 == 0 else "")
    wb.save(path)
    wb.close()


def test_gridbench_recipe_satisfies_sheet_context(tmp_path):
    session = _load_ui("_ui_session")
    path = tmp_path / "gridbench.xlsx"
    _write_gridbench(path)
    workbook, reason = session._open_sheet(path)
    assert reason == ""
    assert workbook is not None


def test_open_sheet_rejects_a_blank_workbook(tmp_path):
    import openpyxl

    session = _load_ui("_ui_session")
    path = tmp_path / "blank.xlsx"
    openpyxl.Workbook().save(path)
    workbook, reason = session._open_sheet(path)
    assert workbook is None
    assert "build_sheet_context" in reason

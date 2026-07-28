"""Local Excel (.xlsx) support for clipgen.

Provides a sheet adapter that matches the gspread Worksheet interface
so spreadsheet.py and clipgen can use local Excel files the same way as Google Sheets.
"""

import datetime as _datetime
import zipfile
from pathlib import Path
from typing import Any, NamedTuple

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

import utils


def _cell_to_str(v: Any) -> str:
    """Convert an openpyxl cell value to a downstream-safe string.

    Time-like cells (``datetime.time``, ``datetime.datetime``,
    ``datetime.timedelta``) are formatted as ``HH:MM:SS`` so timestamp
    parsing downstream behaves the same way it does for Google Sheets,
    which always returns strings. The default ``str()`` on a
    ``datetime.datetime`` prefixes the Excel epoch date (``1899-12-30``)
    and produces unparseable output; this normalizes those away.
    """
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, _datetime.datetime):
        return v.strftime("%H:%M:%S")
    if isinstance(v, _datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, _datetime.timedelta):
        total = int(v.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
    if isinstance(v, float) and v.is_integer():
        # openpyxl often stores whole numbers as floats; gspread returns the
        # displayed value ("5"), not the repr ("5.0").
        return str(int(v))
    return str(v)


class _CellLike(NamedTuple):
    """Minimal cell-like object with .row and .col (1-based) for header lookup."""

    row: int
    col: int


class _SpreadsheetLike(NamedTuple):
    """Minimal spreadsheet-like object with .title and .url = None for Excel."""

    title: str
    url: None = None


class ExcelSheetAdapter:
    """Adapter that makes an openpyxl worksheet look like a gspread Worksheet."""

    def __init__(self, ws: Any, workbook_path: str) -> None:
        self._ws = ws
        self.title = getattr(ws, "title", Path(workbook_path).stem)
        self._workbook_path = workbook_path
        self._data: list[list[str]] = []
        self._load_data()
        self.spreadsheet = _SpreadsheetLike(title=Path(workbook_path).stem)

    def _load_data(self) -> None:
        """Load all cell values into _data as List[List[str]], padded to max column."""
        rows: list[list[str]] = []
        max_col = 0
        for row in self._ws.iter_rows(values_only=True):
            str_row = [_cell_to_str(v) for v in row]
            rows.append(str_row)
            max_col = max(max_col, len(str_row))
        # Pad rows to same length
        for row in rows:
            while len(row) < max_col:
                row.append("")
        self._data = rows

    def find(self, text: str) -> _CellLike | None:
        """Find first cell with exact match. Returns cell-like with .row, .col (1-based)."""
        for row_idx, row in enumerate(self._data):
            for col_idx, cell_value in enumerate(row):
                if cell_value == text:
                    return _CellLike(row=row_idx + 1, col=col_idx + 1)
        return None

    def get_all_values(self) -> list[list[str]]:
        """Return all sheet data as list of rows (list of strings)."""
        return self._data

    def row_values(self, row_1based: int) -> list[str]:
        """Return one row as list of strings. row_1based is 1-based."""
        idx = row_1based - 1
        if 0 <= idx < len(self._data):
            return self._data[idx]
        return []

    @property
    def col_count(self) -> int:
        """Number of columns (max length of any row)."""
        if not self._data:
            return 0
        return max(len(row) for row in self._data)


def _get_worksheet_from_workbook(wb: Any, preferred_name: str | None = None) -> Any:
    """Pick worksheet: preferred_name, else config.WORKSHEET_PRIORITY, else first."""
    sheet_names = wb.sheetnames
    utils.debug_print(f"Available worksheets: {sheet_names}")
    chosen = utils.pick_worksheet_title(sheet_names, preferred_name)
    if chosen is None:
        raise ValueError("Workbook contains no worksheets")
    utils.standard_print(f"Using worksheet: {chosen}")
    return wb[chosen]


def list_worksheet_titles(path: str) -> tuple[list[str], str]:
    """Return ``(titles, recommended)`` for an .xlsx without loading cell data.

    Opens the workbook read-only just for its sheet names, so the Start
    overlay's worksheet dropdown is cheap to populate. ``recommended`` is the
    priority auto-pick (empty string when the workbook has no worksheets).
    """
    path = str(Path(path).resolve())
    if not Path(path).is_file():
        utils.error_print(f"Excel file not found: {path}")
        return [], ""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        titles = list(wb.sheetnames)
        wb.close()
    except (OSError, zipfile.BadZipFile, InvalidFileException) as e:
        utils.error_print(f"Could not read worksheets from {path}: {e}")
        return [], ""
    return titles, utils.pick_worksheet_title(titles) or ""


def open_excel_workbook(
    path: str, worksheet_name: str | None = None
) -> ExcelSheetAdapter | None:
    """Load an .xlsx file and return an ExcelSheetAdapter for the chosen worksheet.

    Args:
        path: Path to the .xlsx file.
        worksheet_name: A worksheet title chosen by the user; falls back to the
            priority auto-pick when unset or absent from the workbook.

    Returns:
        ExcelSheetAdapter, or None on error (prints error).
    """
    path = str(Path(path).resolve())
    if not Path(path).is_file():
        utils.error_print(f"Excel file not found: {path}")
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        ws = _get_worksheet_from_workbook(wb, worksheet_name)
        adapter = ExcelSheetAdapter(ws, path)
        wb.close()
        return adapter
    except (
        KeyError,
        ValueError,
        OSError,
        zipfile.BadZipFile,
        InvalidFileException,
    ) as e:
        # openpyxl raises BadZipFile for a corrupt/non-zip file and
        # InvalidFileException for an unsupported suffix; neither subclasses
        # OSError, so they must be caught explicitly to fail gracefully.
        utils.error_print(f"Could not open Excel file {path}: {e}")
        return None


def list_excel_in_cwd() -> list[str]:
    """Return list of .xlsx file paths in the current working directory (case-insensitive extension)."""
    return sorted(str(p) for p in Path.cwd().iterdir() if p.suffix.lower() == ".xlsx")


def select_excel_file() -> ExcelSheetAdapter | None:
    """Discover .xlsx in cwd: 0 -> error; 1 -> open and return; 2+ -> list and prompt.

    Returns:
        ExcelSheetAdapter on success, None otherwise.
    """
    paths = list_excel_in_cwd()
    if not paths:
        utils.error_print(
            "No .xlsx files found in the current directory.",
            ["Place one or more Excel files (.xlsx) in the working directory."],
        )
        return None
    if len(paths) == 1:
        utils.standard_print(f"Opening Excel file: {Path(paths[0]).name}")
        return open_excel_workbook(paths[0])
    # Multiple files: list and prompt
    if utils.NO_INPUT_MODE:
        utils.error_print(
            "Multiple .xlsx files in current directory; cannot pick one in non-interactive mode.",
            [
                f"Found: {', '.join(Path(p).name for p in paths)}",
                "Pass -s ./path/to/file.xlsx explicitly.",
            ],
        )
        return None
    utils.info_print("Excel files in current directory:")
    for i, p in enumerate(paths, 1):
        utils.info_print(f"  {i}. {Path(p).name}")
    while True:
        choice = utils.read_user_input(
            "\nEnter index (1-based) or filename to open (or Enter to cancel):\n>> "
        ).strip()
        if not choice:
            return None
        # Try as index
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(paths):
                return open_excel_workbook(paths[idx - 1])
            utils.info_print(
                f"Invalid index. Enter a number between 1 and {len(paths)}."
            )
            continue
        # Try as filename
        for p in paths:
            if Path(p).name == choice:
                return open_excel_workbook(p)
        utils.info_print(f'No file named "{choice}". Enter an index or exact filename.')


def _print_credentials_help() -> None:
    """Print troubleshooting steps for setting up Google credentials."""
    utils.info_print(
        "Google Sheets access requires a 'credentials.json' file in the working directory."
    )
    utils.info_print(f"Working directory: {Path.cwd()}")
    utils.info_print("Troubleshooting steps:")
    utils.info_print("  1. Ensure 'credentials.json' exists in the working directory")
    utils.info_print("  2. Verify the credentials file is valid JSON")
    utils.info_print(
        "  3. Check that the service account has access to Google Sheets API"
    )
    utils.info_print(
        "  4. For OAuth flow, delete any existing token files and re-authenticate"
    )


def prompt_for_excel_fallback() -> ExcelSheetAdapter | None:
    """Offer an Excel fallback when Google auth fails.

    Lists .xlsx files in cwd and also accepts a pasted path, 'help' for
    credentials troubleshooting, or an empty input to cancel. Returns the
    opened ExcelSheetAdapter, or None if the user cancelled.
    """
    if utils.NO_INPUT_MODE:
        utils.error_print(
            "Google authentication failed and the Excel fallback prompt is interactive.",
            [
                "Pass -s ./path/to/file.xlsx to use a local Excel file in non-interactive mode."
            ],
        )
        return None
    utils.info_print(
        "No Google credentials available — you can work with a local Excel file instead."
    )
    paths = list_excel_in_cwd()
    if paths:
        utils.info_print("Excel files in current directory:")
        for i, p in enumerate(paths, 1):
            utils.info_print(f"  {i}. {Path(p).name}")
    else:
        utils.info_print("(No .xlsx files found in the current directory.)")

    while True:
        choice = utils.read_user_input(
            "\nEnter an index, filename, path to a .xlsx file, "
            "'help' for credentials setup tips, or Enter to cancel:\n>> "
        ).strip()
        if not choice:
            return None
        if choice.lower() == "help":
            _print_credentials_help()
            continue
        if choice.isdigit() and paths:
            idx = int(choice)
            if 1 <= idx <= len(paths):
                adapter = open_excel_workbook(paths[idx - 1])
                if adapter is not None:
                    return adapter
                continue
            utils.info_print(
                f"Invalid index. Enter a number between 1 and {len(paths)}."
            )
            continue
        # Exact filename match inside cwd.
        matched = False
        for p in paths:
            if Path(p).name == choice:
                adapter = open_excel_workbook(p)
                if adapter is not None:
                    return adapter
                matched = True
                break
        if matched:
            continue
        # Treat as a path (absolute or relative to cwd).
        candidate = Path(choice).expanduser()
        if not candidate.is_absolute():
            candidate = Path.cwd() / candidate
        if candidate.suffix.lower() != ".xlsx":
            utils.info_print(
                "Path must end in .xlsx. Try again, or press Enter to cancel."
            )
            continue
        adapter = open_excel_workbook(str(candidate))
        if adapter is not None:
            return adapter
        # open_excel_workbook already printed the error; loop for retry.
